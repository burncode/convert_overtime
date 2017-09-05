#!/usr/bin/env python
# coding: utf-8
# Description: 转换两张表用的
# Version: 1.1.0
# Date: 20170905


import openpyxl
import os
import datetime


# 函数用于打开表并保存激活表
def openxl(xlsx):
    wb = openpyxl.load_workbook(xlsx)
    sheet = wb.get_active_sheet()
    return sheet

# 转换time的起始时间为17:30
def time_to_1730(time):
    x, y  = time.split('-')
    result = "17:30" + '-' + y
    return result

# 验证日期是否是工作日
def is_weekday(date):
    # 误差需-2
    date = int(date) - 2
    delta = datetime.timedelta(days=date)
    date_1900 = datetime.datetime(1900, 1, 1)
    date = date_1900 + delta
    weekday = date.isoweekday()
    for i in range(1, 6):
        if i == weekday:
            return True
    return False

# 如果是date是工作日, 转换time为17:30开始到结束
def weekday_time(time, date):
    if is_weekday(date):
        return time_to_1730(time)
    else:
        return time

# 读取源表的值,写入目标表中相应位置
def convert(src_sheet):
    keywd = u"加班时长"
    src_data = {}
    for i in range(3, 100):
        date = src_sheet.cell(column=i, row=2).value
        if date == None:
            break
        elif date == keywd:
            # 日期
            date = src_sheet.cell(column=i-1, row=2).value
            for j in range(3, 100):
                # 打印时间
                time = src_sheet.cell(column=i-1, row=j).value
                if time == None:
                    continue
                time = weekday_time(time, date)
                # 打印时长
                overtime = src_sheet.cell(column=i, row=j).value
                if overtime == None:
                    continue
                # 打印姓名
                name = src_sheet.cell(column=1, row=j).value
                src_data.setdefault(date, {})
                src_data[date].setdefault(name, [time, overtime])
    return src_data

# 写入目标xlsx
def writexlsx(src_data, dst_sheet):
    line = 5
    for date in sorted(src_data):
        for name in src_data[date]:
            # print name, date, src_data[date][name][0], src_data[date][name][1]
            dst_sheet['A' + str(line)] = name
            dst_sheet['B' + str(line)] = date
            dst_sheet['C' + str(line)] = src_data[date][name][0]
            dst_sheet['D' + str(line)] = src_data[date][name][1]
            line += 1

# 获取文件夹内的文件路径,只能一个
def get_file_path(path):
    for root, dirs, files in os.walk(path):
        for file in files:
            result = os.path.join(root, file)
            break
        break
    return result

# 主程序
def main():
    src_xlsx = get_file_path('./src')
    src_sheet = openxl(src_xlsx)
    src_date = convert(src_sheet)
    dst_xlsx = get_file_path('./dst')
    wb = openpyxl.load_workbook(dst_xlsx)
    sheet = wb.get_active_sheet()
    writexlsx(src_date, sheet)
    wb.save(dst_xlsx)

if __name__ == '__main__':
    main()
